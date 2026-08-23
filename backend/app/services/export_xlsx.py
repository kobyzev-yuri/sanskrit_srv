"""XLSX export of a Russian translation: page / Sanskrit / Russian."""
from __future__ import annotations

import uuid
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.storage import ensure_dirs
from app.services.translation_rows import translation_rows


def build_project_xlsx(
    project_id: uuid.UUID,
    slug: str,
    title: str,
    pages: list[tuple[int, str]],
    *,
    title_sa: str | None = None,
) -> Path:
    """pages: list of (page_no, translation html)."""
    rows = translation_rows(pages)
    out_dir = ensure_dirs() / "exports" / str(project_id)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{slug}-translation.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Перевод"
    subtitle = (title_sa or "").strip()
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, name="Calibri")
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(size=12, name="Noto Serif Devanagari")
    header_row = 4 if subtitle else 3
    headers = ("Страница", "Санскрит", "Русский")
    fill = PatternFill("solid", fgColor="EFE6D8")
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, name)
        cell.font = Font(bold=True, name="Calibri")
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
    sa_font = Font(name="Noto Serif Devanagari", size=12)
    ru_font = Font(name="Calibri", size=11)
    wrap = Alignment(wrap_text=True, vertical="top")
    if not rows:
        ws.cell(header_row + 1, 1, "Нет страниц с переводом")
    else:
        for i, (page_no, sa, ru) in enumerate(rows, start=header_row + 1):
            c1 = ws.cell(i, 1, page_no)
            c1.alignment = Alignment(vertical="top")
            c2 = ws.cell(i, 2, sa)
            c2.font = sa_font
            c2.alignment = wrap
            c3 = ws.cell(i, 3, ru)
            c3.font = ru_font
            c3.alignment = wrap
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 55
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:C{header_row + max(1, len(rows))}"
    ws.sheet_view.showGridLines = True
    tmp = out_path.with_suffix(f".{uuid.uuid4().hex}.tmp.xlsx")
    wb.save(tmp.as_posix())
    tmp.replace(out_path)
    return out_path
